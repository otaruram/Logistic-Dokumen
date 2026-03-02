"""
Export routes — /api/scans/export-drive, /export-drive-direct
"""

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from config.database import get_db
from models.models import User, Scan
from utils.auth import get_current_active_user

router = APIRouter()


@router.post("/export-drive")
async def export_to_drive(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Export all user scans to Google Drive as Excel.
    Requires the frontend to pass the token — see /export-drive-direct."""
    raise HTTPException(
        status_code=400,
        detail="Please use the 'Export to Drive' button in the app which handles authentication.",
    )


@router.post("/export-drive-direct")
async def export_to_drive_direct(
    token_data: dict,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Export to Drive using token provided by Frontend (Supabase Session).
    Body: { "access_token": "..." }
    """
    import traceback

    try:
        access_token = token_data.get("access_token")
        if not access_token:
            raise HTTPException(status_code=400, detail="Google Access Token required")

        import io
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from services.drive_service import export_to_google_drive_with_token

        # 1. Fetch scans
        scans = (
            db.query(Scan)
            .filter(Scan.user_id == current_user.id)
            .order_by(Scan.created_at.desc())
            .all()
        )
        if not scans:
            raise HTTPException(status_code=404, detail="No records to export")

        data = [
            {
                "No": idx,
                "Date": s.created_at.strftime("%Y-%m-%d %H:%M"),
                "Recipient": s.recipient_name or "-",
                "Extracted Content": s.extracted_text or "-",
                "Status": s.status.upper(),
                "Image Link": s.imagekit_url or "",
                "Signature Link": s.signature_url or "",
            }
            for idx, s in enumerate(scans, 1)
        ]
        df = pd.DataFrame(data)

        # 2. Build styled Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Scan Report")
            ws = writer.sheets["Scan Report"]

            header_fill = PatternFill(
                start_color="111111", end_color="111111", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 25

        output.seek(0)

        # 3. Upload to Drive
        filename = (
            f"Scan_Report_{current_user.email}"
            f"_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"
        )
        print(f"🚀 Starting Drive Export for {current_user.email}...")
        result = export_to_google_drive_with_token(
            access_token, output.read(), filename
        )
        print(f"✅ Drive Export Success: {result}")
        return result

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Export to Drive Direct Error: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
