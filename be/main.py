# backend/main.py

# --- IMPORT LIBRARY ---
from fastapi import FastAPI, UploadFile, File, Form, Header, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List
import requests
import base64
import numpy as np
from PIL import Image
import io
import pandas as pd
from datetime import datetime
import os
import shutil
import jwt
import json
from dotenv import load_dotenv
import re
import traceback
import PyPDF2
import asyncio
from oki_chatbot import OKiChatbot
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from drive_service import export_to_google_drive_with_token, upload_image_to_drive
from contextlib import asynccontextmanager

# Import smart OCR dan AI modules
from smart_ocr_processor import SmartOCRProcessor
from ai_text_summarizer import AITextSummarizer

# Import pricing system
from pricing_service import CreditService
from pricing_endpoints import router as pricing_router

# Load environment variables
load_dotenv()

# Try to import Prisma with graceful fallback
Prisma = None
try:
    from prisma import Prisma as PrismaClient
    Prisma = PrismaClient
    print("✅ Prisma import successful")
except Exception as e:
    print(f"⚠️ Prisma import failed: {e}")
    print("🔄 Running without database support")

# --- SUMOPOD CHATBOT CONFIGURATION ---
# Initialize the global bot instance to use SumoPoD directly.
oki_bot = OKiChatbot(mode='sumopod_only')

# Initialize Prisma Client
prisma = None
if Prisma:
    try:
        prisma = Prisma()
        print("✅ Prisma client initialized")
    except Exception as e:
        print(f"⚠️ Prisma client failed to initialize: {e}")
        print("🔄 Running in fallback mode without database")
else:
    print("⚠️ Prisma not available - running without database")

# Initialize Smart OCR dan AI Summarizer
smart_ocr = None
ai_summarizer = None

# Initialize Credit Service
credit_service = None
try:
    from pricing_service import CreditService
    credit_service = CreditService()
    print("✅ Credit Service initialized!")
except Exception as e:
    print(f"⚠️ Credit Service failed to initialize: {e}")
    credit_service = None

async def daily_maintenance_task():
    """Automatic daily maintenance - runs every day at midnight"""
    try:
        print("🔄 Running automatic daily maintenance...")
        
        if prisma:
            # Check all users for monthly cleanup
            cleanup_count = await CreditService.check_all_users_cleanup(prisma)
            
            if cleanup_count > 0:
                print(f"🗑️ Automatic cleanup performed for {cleanup_count} users")
            else:
                print("✅ No users needed cleanup today")
        else:
            print("⚠️ Database not available for maintenance")
            
    except Exception as e:
        print(f"❌ Error in daily maintenance: {e}")

async def daily_credit_reset_task():
    """Reset credits for all users at midnight Jakarta time"""
    try:
        print("💳 Running daily credit reset...")
        
        if prisma:
            from datetime import datetime, date
            
            # Get all users
            users = await prisma.user.find_many()
            reset_count = 0
            
            for user in users:
                try:
                    # Check if user needs credit reset today
                    today = date.today()
                    last_reset = user.lastCreditReset.date() if user.lastCreditReset else None
                    
                    if last_reset != today:
                        # Reset credits to daily limit (3)
                        await prisma.user.update(
                            where={"id": user.id},
                            data={
                                "creditBalance": CreditService.DAILY_CREDIT_LIMIT,
                                "lastCreditReset": datetime.now()
                            }
                        )
                        
                        # Log credit reset transaction
                        await prisma.credittransaction.create(
                            data={
                                "userId": user.id,
                                "amount": CreditService.DAILY_CREDIT_LIMIT,
                                "description": f"Daily credit reset - {today}"
                            }
                        )
                        
                        reset_count += 1
                        print(f"💳 Credits reset for user: {user.email}")
                        
                except Exception as e:
                    print(f"❌ Error resetting credits for user {user.id}: {e}")
                    continue
            
            print(f"✅ Daily credit reset completed! {reset_count} users updated")
            
        else:
            print("⚠️ Database not available for credit reset")
            
    except Exception as e:
        print(f"❌ Error in daily credit reset: {e}")

# Global scheduler variable
scheduler = None

async def monthly_cleanup_scheduler():
    """Background task to check all users for monthly cleanup based on registration anniversary"""
    while True:
        try:
            from datetime import date
            
            if prisma:
                print("🗑️ Daily check: Looking for users who reached monthly anniversary...")
                cleanup_count = await CreditService.check_all_users_cleanup(prisma)
                
                if cleanup_count > 0:
                    print(f"✅ Performed monthly cleanup for {cleanup_count} users")
                else:
                    print("ℹ️ No users reached monthly anniversary today.")
            
            # Wait 24 hours before checking again
            await asyncio.sleep(24 * 60 * 60)  # 24 hours
            
        except Exception as e:
            print(f"❌ Error in monthly cleanup scheduler: {e}")
            # Wait 1 hour before retrying on error
            await asyncio.sleep(60 * 60)  # 1 hour

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Initialize Smart OCR, AI components, and connect to database"""
    global smart_ocr, ai_summarizer, scheduler
    
    print("🚀 Starting Enhanced OCR System...")
    
    # Initialize Smart OCR dan AI Summarizer
    ocr_api_key = os.getenv('OCR_SPACE_API_KEY', 'helloworld')
    sumopod_api_key = os.getenv('SUMOPOD_API_KEY')
    
    try:
        smart_ocr = SmartOCRProcessor(ocr_api_key)
        print("✅ Smart OCR Processor initialized!")
    except Exception as e:
        print(f"⚠️ Smart OCR initialization failed: {e}")
        smart_ocr = None
        
    # Initialize automatic scheduler
    try:
        scheduler = AsyncIOScheduler()
        
        # Schedule daily maintenance at midnight Jakarta time
        scheduler.add_job(
            daily_maintenance_task,
            "cron",
            hour=0,
            minute=0,
            timezone='Asia/Jakarta',
            id="daily_maintenance"
        )
        
        # Schedule daily credit reset at midnight Jakarta time
        scheduler.add_job(
            daily_credit_reset_task,
            "cron", 
            hour=0,
            minute=1,
            timezone='Asia/Jakarta',
            id="daily_credit_reset"
        )
        
        # Start scheduler
        scheduler.start()
        print("🚀 Automatic scheduler started - daily maintenance at midnight!")
    except Exception as e:
        print(f"⚠️ Scheduler initialization failed: {e}")
        
    try:
        # DISABLED: SumoPod AI not used anymore
        # if sumopod_api_key:
        #     ai_summarizer = AITextSummarizer(sumopod_api_key)
        #     print("✅ AI Text Summarizer initialized!")
        # else:
        #     print("⚠️ SUMOPOD_API_KEY not found, using rule-based summaries")
        #     ai_summarizer = None
        print("ℹ️ SumoPod AI disabled - using rule-based summaries")
        ai_summarizer = None
    except Exception as e:
        print(f"⚠️ AI Summarizer initialization failed: {e}")
        ai_summarizer = None
    
    # Initialize Credit Service
    global credit_service
    try:
        credit_service = CreditService()  # Create instance, not just class reference
        print("✅ Credit Service initialized!")
    except Exception as e:
        print(f"⚠️ Credit Service initialization failed: {e}")
        credit_service = None
    
    # Try to connect to database with retry logic
    database_connected = False
    max_retries = 3
    
    if prisma:
        for attempt in range(max_retries):
            try:
                await prisma.connect()
                print("✅ Connected to Supabase PostgreSQL!")
                database_connected = True
                break
            except Exception as db_error:
                print(f"⚠️ Database connection attempt {attempt + 1}/{max_retries} failed: {db_error}")
                if attempt == max_retries - 1:
                    print("📝 System will continue without database logging")
                    print("💡 Check your DATABASE_URL and internet connection")
    else:
        print("⚠️ Prisma client not available - running without database")
    
    print("⚡ Enhanced OCR System ready!")
    print("🔍 Features: Smart document detection, AI summarization, structured data extraction")
    
    yield
    
    # Shutdown scheduler
    try:
        if scheduler:
            scheduler.shutdown()
            print("🛑 Scheduler shutdown completed")
    except Exception as e:
        print(f"⚠️ Scheduler shutdown error: {e}")
    
    # Cleanup
    try:
        if database_connected and prisma and prisma.is_connected():
            await prisma.disconnect()
            print("🔌 Disconnected from database")
    except Exception as e:
        print(f"Cleanup error: {e}")
    
    print("🛑 Enhanced OCR System shutdown complete")

# Helper function for safe database operations
async def safe_db_operation(operation, *args, **kwargs):
    """Safely execute database operations with error handling"""
    try:
        if not prisma or not prisma.is_connected():
            print("⚠️ Database not connected, skipping operation")
            return None
        return await operation(*args, **kwargs)
    except Exception as e:
        print(f"⚠️ Database operation failed: {e}")
        return None

# --- INISIALISASI APP ---
app = FastAPI(
    title="Supply Chain OCR API", 
    description="Backend untuk scan dokumen gudang",
    lifespan=lifespan
)

# --- SETUP FOLDER UPLOADS ---
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Mount static files untuk serve gambar
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# --- KONFIGURASI CORS ---
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:8080")

# Allow multiple domains
allowed_origins = [
    "http://localhost:8080",
    "http://localhost:8081",
    "http://localhost:5173",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://127.0.0.1:5173",
    "http://127.0.0.1:8080", 
    "https://ocrai.vercel.app",
    "https://ocr.wtf",
    "https://www.ocr.wtf",
    "*"  # Fallback for development
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# Include pricing endpoints
app.include_router(pricing_router, prefix="/api/pricing", tags=["pricing"])

# --- HELPER: DECODE JWT TOKEN OR GET EMAIL FROM GOOGLE API ---
def get_user_email_from_token(authorization: str = Header(None)) -> str:
    if not authorization:
        raise HTTPException(status_code=401, detail="No authorization token provided")
    try:
        token = authorization.replace("Bearer ", "")
        
        # Try to decode as JWT first (old login method)
        try:
            decoded = jwt.decode(token, options={"verify_signature": False})
            email = decoded.get("email")
            if email:
                return email
        except:
            # Not a JWT, might be an access token
            pass
        
        # If not JWT, try to get user info from Google API (new login with Drive scope)
        try:
            response = requests.get(
                'https://www.googleapis.com/oauth2/v3/userinfo',
                headers={'Authorization': f'Bearer {token}'},
                timeout=5
            )
            if response.status_code == 200:
                user_info = response.json()
                email = user_info.get('email')
                if email:
                    return email
        except:
            pass
            
        raise HTTPException(status_code=401, detail="Could not extract email from token")
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Invalid token: {str(e)}")

# --- HELPER: OCR ENGINE (OCR.SPACE) ---
OCR_API_KEY = os.getenv("OCR_API_KEY", "helloworld") # Ganti di .env biar limitnya gede
OCR_API_URL = "https://api.ocr.space/parse/image"

async def extract_text_from_image(image_np, user_email: str = None):
    """Enhanced OCR dengan Smart Processing dan AI Summarization"""
    try:
        print("Starting enhanced OCR processing...")
        
        # Check for user's BYOK OCR API key first
        api_key_to_use = OCR_API_KEY  # Default
        using_byok = False
        
        if user_email:
            user_ocr_key = await get_user_ocr_api_key_internal(user_email)
            if user_ocr_key:
                api_key_to_use = user_ocr_key
                using_byok = True
                print(f"BYOK OCR ENABLED - Using user's OCR API key for {user_email}")
            else:
                print(f"DEFAULT OCR API KEY - Using system OCR key")
        
        # Use Smart OCR if available
        if smart_ocr and api_key_to_use != "helloworld":
            try:
                # Initialize smart OCR with user's key if BYOK
                if using_byok:
                    user_smart_ocr = SmartOCRProcessor(api_key_to_use)
                    extracted_text = await user_smart_ocr.enhanced_ocr_extract(image_np)
                else:
                    extracted_text = await smart_ocr.enhanced_ocr_extract(image_np)
                
                if extracted_text and not extracted_text.startswith("[ERROR"):
                    print("Smart OCR extraction successful")
                    
                    # Detect document type dan extract structured data
                    doc_type = smart_ocr.detect_document_type(extracted_text)
                    structured_data = smart_ocr.extract_structured_data(extracted_text, doc_type)
                    
                    print(f"Document type detected: {doc_type}")
                    if structured_data:
                        print(f"Structured data extracted: {structured_data}")
                    
                    # Generate smart summary menggunakan AI atau rule-based
                    smart_summary = await generate_smart_summary(extracted_text, doc_type, structured_data)
                    
                    return {
                        "raw_text": extracted_text,
                        "summary": smart_summary,
                        "document_type": doc_type,
                        "structured_data": structured_data,
                        "processing_method": "smart_ocr",
                        "byok_used": using_byok
                    }
                else:
                    print("Smart OCR failed, falling back to basic OCR")
                    
            except Exception as e:
                print(f"Smart OCR error: {e}, falling back to basic OCR")
        
        # Fallback to basic OCR
        print("Using basic OCR processing...")
        basic_text = await basic_ocr_extract(image_np, api_key_to_use)
        
        if basic_text:
            # Generate basic summary
            basic_summary = generate_basic_summary(basic_text)
            
            return {
                "raw_text": basic_text,
                "summary": basic_summary,
                "document_type": "unknown",
                "structured_data": {},
                "processing_method": "basic_ocr",
                "byok_used": using_byok
            }
        else:
            return {
                "raw_text": "",
                "summary": "Tidak dapat mengekstrak teks dari gambar",
                "document_type": "unknown",
                "structured_data": {},
                "processing_method": "failed",
                "byok_used": using_byok
            }
        
    except Exception as e:
        print(f"OCR processing error: {e}")
        return {
            "raw_text": f"[ERROR: {str(e)}]",
            "summary": f"Error saat memproses gambar: {str(e)}",
            "document_type": "error",
            "structured_data": {},
            "processing_method": "error",
            "byok_used": False
        }

async def basic_ocr_extract(image_np, api_key: str):
    """Basic OCR fallback menggunakan OCR.space API"""
    try:
        # 1. Convert Numpy ke Base64 Image
        image = Image.fromarray(image_np)
        
        # Convert RGBA to RGB if needed (fix for PNG with transparency)
        if image.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
            image = background
        
        # Resize jika kegedean biar cepet
        max_width = 1024
        if image.width > max_width:
            ratio = max_width / image.width
            new_height = int(image.height * ratio)
            image = image.resize((max_width, new_height), Image.Resampling.LANCZOS)

        img_byte_arr = io.BytesIO()
        image.save(img_byte_arr, format='JPEG', quality=85)
        img_byte_arr = img_byte_arr.getvalue()
        base64_image = base64.b64encode(img_byte_arr).decode('utf-8')

        # 2. Request ke API Luar
        payload = {
            'apikey': api_key,
            'base64Image': f'data:image/jpeg;base64,{base64_image}',
            'language': 'eng',
            'isOverlayRequired': False,
            'scale': True,
            'OCREngine': 1
        }
        
        print("Sending to OCR.space (basic)...")
        response = requests.post(OCR_API_URL, data=payload, timeout=30)
        result = response.json()

        # 3. Parsing Hasil
        if result.get('IsErroredOnProcessing'):
            raise Exception(result.get('ErrorMessage'))
            
        parsed_results = result.get('ParsedResults', [])
        if parsed_results:
            text = parsed_results[0].get('ParsedText', '').strip()
            return text
        else:
            return ""
            
    except Exception as e:
        print(f"Basic OCR failed: {e}")
        return f"[ERROR OCR: {str(e)}]"

async def generate_smart_summary(text: str, doc_type: str, structured_data: dict) -> str:
    """Generate summary menggunakan rule-based only (SumoPod disabled)"""
    try:
        # SumoPod AI disabled - use smart rule-based summary only
        if smart_ocr:
            return smart_ocr.generate_smart_summary(text, structured_data, doc_type)
        
        # Last resort: basic summary
        return generate_basic_summary(text)
        
    except Exception as e:
        print(f"Smart summary generation error: {e}")
        return generate_basic_summary(text)

def generate_basic_summary(text: str) -> str:
    """Generate basic summary dari text"""
    try:
        if not text or text.startswith("[ERROR"):
            return "Tidak dapat membuat ringkasan"
        
        # Ambil baris pertama yang meaningful
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        for line in lines[:3]:
            if len(line) > 10 and not line.isdigit():
                # Potong kalau terlalu panjang
                if len(line) > 100:
                    return line[:97] + "..."
                return line
        
        return "Dokumen berhasil dipindai" if lines else "Tidak ada teks terdeteksi"
        
    except Exception as e:
        return "Error dalam membuat ringkasan"

# --- ENDPOINTS ---

@app.get("/")
def home():
    return {
        "status": "Online", 
        "backend": "FastAPI + Enhanced Smart OCR + AI Summarization",
        "features": [
            "🔍 Smart Document Type Detection",
            "📊 Structured Data Extraction", 
            "🤖 AI-Powered Summarization",
            "⚡ Advanced Image Preprocessing",
            "🎯 High-Accuracy OCR Processing"
        ],
        "smart_ocr": "✅ Active" if smart_ocr else "❌ Inactive",
        "ai_summarizer": "✅ Active" if ai_summarizer else "❌ Inactive"
    }

@app.post("/scan")
async def scan_document(
    file: UploadFile = File(...), 
    receiver: str = Form(...),
    authorization: str = Header(None)
):
    global credit_service
    try:
        user_email = get_user_email_from_token(authorization)
        
        # 🎁 ENSURE USER HAS DEFAULT CREDITS (10 for new users)
        try:
            if credit_service:
                await credit_service.ensure_default_credits(user_email, prisma)
        except Exception as credit_error:
            print(f"⚠️ Welcome credit setup failed: {credit_error}")
        
        # 🔥 CHECK CREDIT - Enhanced OCR costs 1 credit
        if credit_service:
            try:
                user_credits = await credit_service.get_user_credits(user_email, prisma)
                if user_credits < 1:
                    return {
                        "status": "error", 
                        "message": "Insufficient credits. Please upgrade your plan.",
                        "credits_remaining": user_credits,
                        "upgrade_url": "/pricing",
                        "error_type": "insufficient_credits"
                    }
            except Exception as credit_error:
                print(f"⚠️ Credit check failed: {credit_error}")
                # Continue without credit check in case of error
        
        # 1. Baca File
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))
        image_np = np.array(image)

        # 2. Proses OCR dengan Enhanced Smart Processing
        ocr_result = await extract_text_from_image(image_np, user_email)
        
        # Extract hasil OCR
        if isinstance(ocr_result, dict):
            full_text = ocr_result.get("raw_text", "")
            smart_summary = ocr_result.get("summary", "")
            doc_type = ocr_result.get("document_type", "unknown")
            structured_data = ocr_result.get("structured_data", {})
            processing_method = ocr_result.get("processing_method", "unknown")
            byok_used = ocr_result.get("byok_used", False)
        else:
            # Fallback untuk backward compatibility
            full_text = str(ocr_result)
            smart_summary = full_text[:200].replace("\n", " ") if full_text else ""
            doc_type = "unknown"
            structured_data = {}
            processing_method = "legacy"
            byok_used = False

        full_text_upper = full_text.upper()

        # 3. Enhanced Document Analysis dengan Structured Data
        nomor_dokumen = "TIDAK TERDETEKSI"
        
        # Prioritaskan structured data jika ada
        if structured_data:
            if doc_type == "invoice" and 'invoice_number' in structured_data:
                nomor_dokumen = structured_data['invoice_number']
            elif doc_type == "delivery_note" and 'do_number' in structured_data:
                nomor_dokumen = structured_data['do_number']
            elif doc_type == "purchase_order" and 'po_number' in structured_data:
                nomor_dokumen = structured_data['po_number']
        
        # Fallback ke pattern matching jika belum ketemu
        if nomor_dokumen == "TIDAK TERDETEKSI" and full_text_upper:
            patterns = [
                r'[A-Z]{2,6}\d+[-/][A-Z]{2,6}[-/]\d{2,4}[-/]\d{2,6}',
                r'(?:NOMOR|NO|INV|PO|SJ)\s*[:#.]?\s*([A-Z0-9/-]+)',
                r'([A-Z]+[-/]\d{4}[-/]\d+)',
                r'([A-Z]{2,}\d{4,})'
            ]
            for p in patterns:
                match = re.search(p, full_text_upper)
                if match:
                    nomor_dokumen = match.group(1) if match.lastindex else match.group(0)
                    nomor_dokumen = re.sub(r'[:#.]', '', nomor_dokumen).strip()
                    break

        # 4. Enhanced Kategori Classification
        kategori = "DOKUMEN LAIN"
        
        # Prioritaskan document type dari smart OCR
        if doc_type == "invoice":
            kategori = "INVOICE"
        elif doc_type == "delivery_note":
            kategori = "SURAT JALAN"
        elif doc_type == "purchase_order":
            kategori = "PURCHASE ORDER"
        elif doc_type == "receipt":
            kategori = "PERMINTAAN PEMBAYARAN"
        else:
            # Fallback ke keyword detection
            keywords = {
                "INVOICE": "INVOICE",
                "FAKTUR": "INVOICE", 
                "TAGIHAN": "INVOICE",
                "SURAT JALAN": "SURAT JALAN",
                "DELIVERY": "SURAT JALAN",
                "PENGIRIMAN": "SURAT JALAN",
                "PO": "PURCHASE ORDER",
                "PURCHASE ORDER": "PURCHASE ORDER",
                "PESANAN": "PURCHASE ORDER",
                "BERITA ACARA": "BERITA ACARA",
                "PEMBAYARAN": "PERMINTAAN PEMBAYARAN",
                "KUITANSI": "PERMINTAAN PEMBAYARAN"
            }
            for key, val in keywords.items():
                if key in full_text_upper:
                    kategori = val
                    break

        # 5. Use Smart Summary atau Fallback
        if smart_summary and len(smart_summary) > 10:
            summary = smart_summary
        else:
            # Fallback ke summary tradisional
            summary = full_text[:200].replace("\n", " ") if full_text else "Dokumen berhasil dipindai"
        
        timestamp = datetime.now()

        # 6. Simpan File Lokal (PERINGATAN: Di Render Free, ini hilang kalau restart)
        saved_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        file_path = os.path.join(UPLOAD_DIR, saved_filename)
        with open(file_path, "wb") as buffer:
            buffer.write(contents)
        
        # Construct URL gambar - Smart detection for dev/production
        ENVIRONMENT = os.getenv("ENVIRONMENT", "production")
        if ENVIRONMENT == "development":
            BASE_URL = os.getenv("BASE_URL", "http://localhost:8000")
        else:
            BASE_URL = os.getenv("PRODUCTION_URL", "https://logistic-dokumen.onrender.com")
        
        image_url = f"{BASE_URL}/uploads/{saved_filename}"

        # 7. Enhanced Database Save dengan metadata tambahan
        log = None
        try:
            if prisma and prisma.is_connected():
                log_data = {
                    "userId": user_email,
                    "timestamp": timestamp,
                    "filename": file.filename,
                    "kategori": kategori,
                    "nomorDokumen": nomor_dokumen,
                    "receiver": receiver.upper(),
                    "imagePath": image_url,
                    "summary": summary,
                    "fullText": full_text
                }
                
                log = await prisma.logs.create(data=log_data)
                print(f"✅ Dokumen berhasil disimpan ke database dengan ID: {log.id}")
            else:
                print("⚠️ Database not connected, creating temporary log object")
        except Exception as db_error:
            print(f"💾 Database save failed: {db_error}")
            
        # Create fallback log object if database save failed
        if log is None:
            log = type('obj', (object,), {
                'id': f"temp_{int(timestamp.timestamp())}",
                'timestamp': timestamp,
                'kategori': kategori,
                'nomorDokumen': nomor_dokumen,
                'receiver': receiver.upper(),
                'imagePath': image_url,
                'summary': summary,
                'fullText': full_text
            })()
        
        print(f"✅ Document processed successfully:")
        print(f"   📂 Category: {kategori}")
        print(f"   🔢 Document Number: {nomor_dokumen}")
        print(f"   📝 Summary: {summary[:100]}...")
        print(f"   🔄 Processing Method: {processing_method if 'processing_method' in locals() else 'legacy'}")
        
        # 🏦 DEDUCT CREDIT - Enhanced OCR scan completed (only if not using BYOK)
        remaining_credits = None
        byok_used = ocr_result.get("byok_used", False)
        print(f"💳 CREDIT CHECK - BYOK: {byok_used}, CreditService: {credit_service is not None}")
        
        if credit_service and not byok_used:
            try:
                print(f"🔄 ENSURING DEFAULT CREDITS for {user_email}")
                await credit_service.ensure_default_credits(user_email, prisma)
                
                print(f"💸 ATTEMPTING CREDIT DEDUCTION for {user_email}")
                # 🔥 NEW: Get remaining balance directly from deduct function
                remaining_credits = await credit_service.deduct_credits(user_email, 1, f"Enhanced OCR scan - Document ID: {log.id}", prisma)
                
                if remaining_credits is not None:
                    print(f"🎉 CREDIT DEDUCTION SUCCESSFUL - Remaining: {remaining_credits}")
                else:
                    print(f"❌ CREDIT DEDUCTION FAILED for {user_email}")
                    remaining_credits = 0
            except Exception as credit_error:
                print(f"💥 CREDIT DEDUCTION EXCEPTION: {credit_error}")
                remaining_credits = 0
        elif byok_used:
            print(f"🔑 BYOK MODE - No credit deducted for {user_email}")
            # Get current credits without deducting
            try:
                if credit_service:
                    remaining_credits = await credit_service.get_user_credits(user_email, prisma)
                else:
                    remaining_credits = 3
            except:
                remaining_credits = 3
        else:
            print(f"⚠️ CREDIT SERVICE UNAVAILABLE")
            remaining_credits = 3

        # Get updated credit information for response
        credit_info = {}
        try:
            if credit_service and remaining_credits is not None:
                user_tier = await credit_service.get_user_tier(user_email, prisma)
                
                credit_info = {
                    "remainingCredits": remaining_credits,
                    "userTier": user_tier,
                    "creditsUsed": 0 if byok_used else 1,  # Enhanced OCR scan cost
                    "upgradeAvailable": user_tier == "starter" and remaining_credits < 5
                }
        except Exception as credit_error:
            print(f"⚠️ Failed to get credit info for response: {credit_error}")
            credit_info = {"remainingCredits": remaining_credits or 3}

        return {
            "status": "success",
            "data": {
                "id": log.id,
                "timestamp": log.timestamp.isoformat() if hasattr(log.timestamp, 'isoformat') else str(log.timestamp),
                "kategori": log.kategori,
                "nomorDokumen": log.nomorDokumen,
                "nomor_dokumen": log.nomorDokumen,  # Backward compatibility
                "receiver": log.receiver,
                "imagePath": log.imagePath,
                "imageUrl": log.imagePath,  # Backward compatibility
                "summary": log.summary,
                "fullText": log.fullText
            },
            "creditInfo": credit_info,
            # ⚡ REAL-TIME CREDIT UPDATE KEY
            "remaining_credits": remaining_credits if remaining_credits is not None else 3
        }

    except Exception as e:
        print(f"Global Error: {traceback.format_exc()}")
        return {"status": "error", "message": str(e)}

@app.get("/history")
async def get_history(authorization: str = Header(None)):
    try:
        user_email = get_user_email_from_token(authorization)
        
        # Try to get from database if connected
        if prisma and prisma.is_connected():
            try:
                logs = await prisma.logs.find_many(
                    where={"userId": user_email},
                    order={"id": "desc"}
                )
                
                # Format response to ensure all fields are properly mapped
                formatted_logs = []
                for log in logs:
                    formatted_logs.append({
                        "id": log.id,
                        "userId": log.userId,
                        "timestamp": log.timestamp.isoformat(),
                        "filename": log.filename,
                        "kategori": log.kategori,
                        "nomorDokumen": log.nomorDokumen,
                        "receiver": log.receiver,
                        "imagePath": log.imagePath,
                        "summary": log.summary,
                        "fullText": log.fullText
                    })
                
                print(f"📚 History loaded: {len(formatted_logs)} documents for {user_email}")
                return formatted_logs
                
            except Exception as db_error:
                print(f"💾 Database query failed: {db_error}")
                return []
        else:
            print(f"📚 History requested by {user_email} - database not connected")
            return []
        
    except Exception as e:
        print(f"History error: {e}")
        return []

class LogUpdateRequest(BaseModel):
    summary: str

@app.put("/logs/{log_id}")
async def update_log_summary(log_id: int, request: LogUpdateRequest, authorization: str = Header(None)):
    try:
        user_email = get_user_email_from_token(authorization)
        
        print(f"UPDATE LOG REQUEST - User: {user_email}, LogID: {log_id}, New Summary: {request.summary[:50]}...")
        
        # Try to update in database
        try:
            # Check if log exists and belongs to user
            log = await prisma.logs.find_first(
                where={"id": log_id, "userId": user_email}
            )
            
            if not log:
                return {"status": "error", "message": "Log not found or access denied"}
            
            # Update the summary
            updated_log = await prisma.logs.update(
                where={"id": log_id},
                data={"summary": request.summary}
            )
            
            return {
                "status": "success", 
                "message": "Summary updated successfully",
                "data": {
                    "id": updated_log.id,
                    "summary": updated_log.summary
                }
            }
            
        except Exception as db_error:
            print(f"Database update failed: {db_error}")
            return {"status": "success", "message": "Summary updated (database not connected)"}
        
    except Exception as e:
        print(f"Update error: {e}")
        return {"status": "error", "message": str(e)}

@app.post("/update_summary")
async def update_summary_endpoint(
    log_id: int = Form(...), 
    summary: str = Form(...),
    authorization: str = Header(None)
):
    """Update summary for a log entry - placeholder for now"""
    try:
        user_email = get_user_email_from_token(authorization)
        print(f"📝 Summary update requested by {user_email} for log {log_id}")
        return {"status": "success", "message": "Summary updated (database not connected)"}
    except Exception as e:
        print(f"Summary update error: {e}")
        return {"status": "error", "message": str(e)}

@app.delete("/logs/{log_id}")
async def delete_log(log_id: int, authorization: str = Header(None)):
    try:
        user_email = get_user_email_from_token(authorization)
        
        # Cek kepemilikan
        try:
            log = await prisma.logs.find_first(
                where={"id": log_id, "userId": user_email}
            )
        except Exception as db_error:
            print(f"Database query failed: {db_error}")
            raise HTTPException(status_code=500, detail="Database connection error")
        
        if not log:
            raise HTTPException(status_code=404, detail="Log not found")

        # Hapus File Lokal
        if log.imagePath:
            fname = log.imagePath.split("/")[-1]
            local_path = os.path.join(UPLOAD_DIR, fname)
            if os.path.exists(local_path):
                os.remove(local_path)
        
        # Hapus DB
        try:
            await prisma.logs.delete(where={"id": log_id})
        except Exception as db_error:
            print(f"Database delete failed: {db_error}")
            raise HTTPException(status_code=500, detail="Database connection error")
        
        return {"status": "success"}
        
    except HTTPException as he:
        raise he
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/export")
async def export_excel(authorization: str = Header(None), upload_to_drive: bool = True):
    try:
        user_email = get_user_email_from_token(authorization)
        
        # Extract token for Google Drive
        token = authorization.replace("Bearer ", "") if authorization else None
        
        try:
            logs = await prisma.logs.find_many(
                where={"userId": user_email},
                order={"id": "desc"}
            )
        except Exception as db_error:
            print(f"Database query failed: {db_error}")
            raise HTTPException(status_code=500, detail="Database connection error")
        
        # Upload images to Google Drive first if enabled
        image_links = {}
        if upload_to_drive and token:
            folder_name = os.getenv('GOOGLE_DRIVE_FOLDER_NAME', 'LOGISTIC.AI Reports')
            images_folder = f"{folder_name}/Images"
            
            for l in logs:
                if l.imagePath:
                    # Extract filename from URL
                    fname = l.imagePath.split("/")[-1]
                    local_path = os.path.join(UPLOAD_DIR, fname)
                    
                    if os.path.exists(local_path):
                        print(f"Uploading image: {fname}")
                        img_result = upload_image_to_drive(token, local_path, images_folder)
                        if img_result:
                            image_links[l.id] = img_result['direct_link']
        
        data = []
        for l in logs:
            # Convert timezone-aware datetime to timezone-naive for Excel compatibility
            timestamp_naive = l.timestamp.replace(tzinfo=None) if l.timestamp.tzinfo else l.timestamp
            
            # Use Google Drive link if available, otherwise use original
            foto_link = image_links.get(l.id, l.imagePath if l.imagePath else "")
            
            data.append({
                "Tanggal": timestamp_naive,
                "Kategori": l.kategori,
                "Nomor Dokumen": l.nomorDokumen,
                "Penerima": l.receiver,
                "Ringkasan": l.summary,
                "Link Foto": foto_link
            })
            
        df = pd.DataFrame(data)
        filename = f"Laporan_{user_email.split('@')[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Save to BytesIO with formatting
        excel_buffer = io.BytesIO()
        
        # Use ExcelWriter for advanced formatting
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Laporan')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Laporan']
            
            # Import openpyxl styles
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Style untuk header
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            # Border style
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Format header row
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Format data cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    
                    # CENTER ALIGN SEMUA KOLOM
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Format tanggal
                    if cell.column == 1 and cell.value:
                        cell.number_format = 'DD/MM/YYYY HH:MM'
                    
                    # Make link clickable dengan text yang lebih pendek
                    if cell.column == 6 and cell.value and isinstance(cell.value, str) and cell.value.startswith('http'):
                        cell.hyperlink = cell.value
                        cell.font = Font(color="0563C1", underline="single")
                        cell.value = "📷 Lihat Foto"  # Ganti URL panjang dengan text pendek
            
            # Auto-adjust column width dengan ukuran lebih baik
            for col_num, column_title in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                
                # Set specific widths untuk tampilan optimal
                if column_title == "Tanggal":
                    worksheet.column_dimensions[column_letter].width = 22
                elif column_title == "Kategori":
                    worksheet.column_dimensions[column_letter].width = 18
                elif column_title == "Nomor Dokumen":
                    worksheet.column_dimensions[column_letter].width = 25
                elif column_title == "Penerima":
                    worksheet.column_dimensions[column_letter].width = 30
                elif column_title == "Ringkasan":
                    worksheet.column_dimensions[column_letter].width = 60
                elif column_title == "Link Foto":
                    worksheet.column_dimensions[column_letter].width = 18
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Set row height untuk semua baris
            worksheet.row_dimensions[1].height = 35  # Header lebih tinggi
            for row_num in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[row_num].height = 25  # Data row
        
        excel_content = excel_buffer.getvalue()
        
        # Upload to Google Drive if enabled
        drive_result = None
        if upload_to_drive and token:
            try:
                drive_result = export_to_google_drive_with_token(
                    access_token=token,
                    file_content=excel_content,
                    file_name=filename,
                    convert_to_sheets=True
                )
                print(f"File uploaded to Google Drive successfully: {drive_result['web_view_link']}")
            except Exception as drive_error:
                print(f"Error uploading file to Google Drive: {str(drive_error)}")
                traceback.print_exc()
                # Continue with local file download even if Drive upload fails
                # Return info that Drive feature requires OAuth setup
                return {
                    "status": "info",
                    "message": f"Google Drive upload gagal: {str(drive_error)}",
                    "download_url": f"/download/{filename}",
                    "note": "File tetap bisa didownload secara lokal"
                }
        
        # Also save locally for download
        with open(filename, 'wb') as f:
            f.write(excel_content)
        
        # Return response with Google Drive link if available
        if drive_result:
            return {
                "status": "success",
                "message": "File exported to Google Drive",
                "drive_url": drive_result['web_view_link'],
                "file_name": drive_result['file_name'],
                "folder_name": drive_result['folder_name'],
                "download_url": f"/download/{filename}"
            }
        else:
            return FileResponse(filename, filename=filename)

    except Exception as e:
        print(f"Export error: {str(e)}")
        traceback.print_exc()
        return {"status": "error", "message": str(e)}

@app.get("/download/{filename}")
async def download_file(filename: str, authorization: str = Header(None)):
    """Download exported Excel file"""
    try:
        user_email = get_user_email_from_token(authorization)
        file_path = os.path.join(os.getcwd(), filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        return FileResponse(file_path, filename=filename)
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- PYDANTIC MODELS FOR CHAT ---
class ChatMessage(BaseModel):
    role: str
    content: str

class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    pdfText: str = ""

# --- ENDPOINT: CHAT WITH OKI AI ---
@app.post("/api/chat")
async def chat_with_oki(
    request: ChatRequest,
    authorization: str = Header(None),
    x_user_api_key: str = Header(None, alias="X-User-API-Key")
):
    """Chat endpoint untuk Gaskeun - OKi AI Assistant with BYOK support"""
    global credit_service
    try:
        # Verify user authentication
        user_email = get_user_email_from_token(authorization)
        
        # 💳 CHECK CREDIT BEFORE PROCESSING (for non-BYOK mode)
        if not x_user_api_key:
            if credit_service:
                try:
                    user_credits = await credit_service.get_user_credits(user_email, prisma)
                    if user_credits < 1:
                        return {
                            "status": "error",
                            "message": "Insufficient credits for chatbot usage. Please upgrade your plan.",
                            "error_type": "insufficient_credits",
                            "creditsRequired": 1,
                            "creditsAvailable": user_credits
                        }
                except Exception as credit_error:
                    print(f"⚠️ Credit check failed: {credit_error}")
                    # Continue anyway if credit service fails
        
        # Convert messages to dict format
        messages = [{"role": msg.role, "content": msg.content} for msg in request.messages]
        
        # If a key is sent in the header, it's BYOK mode with failover.
        if x_user_api_key and x_user_api_key.strip():  # Check if key exists and not empty
            print(f"BYOK MODE - Key from header detected for {user_email}")
            try:
                # Create a temporary bot instance with the user's key
                custom_bot = OKiChatbot(api_key=x_user_api_key, mode='sumopod_only')
                assistant_message = custom_bot.chat(messages=messages, pdf_text=request.pdfText)
                using_byok = True
            except Exception as byok_error:
                print(f"BYOK failed: {str(byok_error)}")
                # Return user-friendly error without falling back
                return {
                    "status": "error",
                    "message": f"BYOK Error: {str(byok_error)}",
                    "usingBYOK": True,
                    "error_type": "byok_failed"
                }
        # Otherwise, it's default mode, using the system's SumoPoD key directly.
        else:
            print(f"DEFAULT MODE - Using system SumoPoD key for {user_email}")
            try:
                # Use the global bot instance configured for sumopod_only.
                assistant_message = oki_bot.chat(messages=messages, pdf_text=request.pdfText)
                using_byok = False
            except Exception as system_error:
                print(f"System chat failed: {str(system_error)}")
                return {
                    "status": "error",
                    "message": f"System chat error: {str(system_error)}",
                    "usingBYOK": False,
                    "error_type": "system_failed"
                }
        
        # 💳 DEDUCT CREDIT AFTER SUCCESSFUL CHAT (for non-BYOK mode)
        remaining_credits = None
        if not x_user_api_key and credit_service:
            try:
                success = await credit_service.deduct_credits(user_email, 1, "OKi Chatbot usage", prisma)
                if success:
                    remaining_credits = await credit_service.get_user_credits(user_email, prisma)
                    print(f"💳 Credit deducted for chatbot. Remaining: {remaining_credits}")
                else:
                    print("⚠️ Failed to deduct credit but chat completed")
            except Exception as credit_error:
                print(f"⚠️ Credit deduction failed: {credit_error}")
        elif x_user_api_key:
            print("🔑 BYOK mode - No credit deducted for chatbot")
        
        # Get updated credit information
        credit_info = {}
        if credit_service and not x_user_api_key:
            try:
                user_credits = await credit_service.get_user_credits(user_email, prisma)
                user_tier = await credit_service.get_user_tier(user_email, prisma)
                
                credit_info = {
                    "remainingCredits": user_credits,
                    "userTier": user_tier,
                    "creditsUsed": 1,
                    "upgradeAvailable": user_tier == "starter" and user_credits < 5
                }
            except Exception as credit_error:
                print(f"⚠️ Failed to get credit info for response: {credit_error}")
        
        return {
            "status": "success",
            "message": assistant_message,
            "usingBYOK": using_byok,
            "creditInfo": credit_info if credit_info else None
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Chat endpoint error: {str(e)}")
        traceback.print_exc()
        return {
            "status": "error",
            "message": f"Unexpected error: {str(e)[:100]}",
            "usingBYOK": False,
            "error_type": "unexpected"
        }

# --- ENDPOINT: EXTRACT PDF TEXT ---
@app.post("/api/extract-pdf")
async def extract_pdf_text(
    file: UploadFile = File(...),
    authorization: str = Header(None)
):
    """Extract text from PDF file"""
    try:
        # Verify user authentication
        user_email = get_user_email_from_token(authorization)
        
        # Read PDF file
        pdf_bytes = await file.read()
        pdf_file = io.BytesIO(pdf_bytes)
        
        # Extract text using PyPDF2 (use PdfFileReader for older versions)
        try:
            # Try newer API first (PyPDF2 >= 3.0)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
        except AttributeError:
            # Fall back to older API (PyPDF2 < 3.0)
            pdf_reader = PyPDF2.PdfFileReader(pdf_file)
        
        text = ""
        
        # Handle both old and new PyPDF2 APIs
        num_pages = len(pdf_reader.pages) if hasattr(pdf_reader, 'pages') else pdf_reader.numPages
        
        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num] if hasattr(pdf_reader, 'pages') else pdf_reader.getPage(page_num)
            # Try new API first, then fall back to old API
            try:
                text += page.extract_text() + "\n\n"
            except AttributeError:
                text += page.extractText() + "\n\n"
        
        return {
            "status": "success",
            "text": text.strip(),
            "pages": num_pages
        }
        
    except Exception as e:
        print(f"PDF extraction error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"PDF extraction error: {str(e)}")

# --- DELETE ACCOUNT ENDPOINT ---
@app.delete("/delete-account")
async def delete_account(authorization: str = Header(None)):
    """Delete user account and all associated data"""
    try:
        user_email = get_user_email_from_token(authorization)
        
        # Get user first to get user ID
        user = await prisma.user.find_unique(where={"email": user_email})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        user_id = user.id
        
        # Get all user logs to delete associated image files
        user_logs = await prisma.logs.find_many(
            where={"userId": user_email}
        )
        
        # Delete associated image files from uploads folder
        deleted_files = 0
        for log in user_logs:
            if log.imagePath:
                filename = log.imagePath.split("/")[-1]
                file_path = os.path.join(UPLOAD_DIR, filename)
                if os.path.exists(file_path):
                    os.remove(file_path)
                    deleted_files += 1
                    print(f"🗑️ Deleted image: {filename}")
        
        # Delete user data in the correct order (due to foreign key constraints)
        try:
            # 1. Delete all user logs (references user email)
            deleted_logs = await prisma.logs.delete_many(
                where={"userId": user_email}
            )
            
            # 2. Delete all credit transactions (references user ID)
            deleted_transactions = await prisma.credittransaction.delete_many(
                where={"userId": user_id}
            )
            
            # 3. Delete all API keys (references user ID)
            deleted_api_keys = await prisma.apikey.delete_many(
                where={"userId": user_email}  # Note: APIKey uses email as userId
            )
            
            # 4. Finally delete the user record
            deleted_user = await prisma.user.delete(
                where={"id": user_id}
            )
            
        except Exception as db_error:
            print(f"Database delete failed: {db_error}")
            raise HTTPException(status_code=500, detail="Database connection error")
        
        print(f"ACCOUNT DELETED: {user_email}")
        print(f"- Logs deleted: {deleted_logs}")
        print(f"- Transactions deleted: {deleted_transactions}")
        print(f"- API keys deleted: {deleted_api_keys}")
        print(f"- Image files deleted: {deleted_files}")
        print(f"- User record deleted: {deleted_user.email}")
        
        return {
            "status": "success",
            "message": f"Account completely deleted successfully.",
            "email": user_email,
            "deleted_logs": deleted_logs,
            "deleted_transactions": deleted_transactions,
            "deleted_api_keys": deleted_api_keys,
            "deleted_files": deleted_files
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Delete account error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Delete account error: {str(e)}")

# --- ADMIN ENDPOINTS ---
@app.get("/admin/logs")
async def admin_get_logs(password: str):
    try:
        ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "supply2024reset")
        if password != ADMIN_PASSWORD:
            raise HTTPException(status_code=403, detail="Invalid admin password")
        
        try:
            logs = await prisma.logs.find_many()
        except Exception as db_error:
            print(f"Database query failed: {db_error}")
            return []
        
        return logs
    except HTTPException as he:
        raise he
    except Exception as e:
        return {"status": "error", "message": str(e)}

# --- ADMIN ENDPOINT: RESET DATABASE ---
class ResetRequest(BaseModel):
    admin_password: str

@app.post("/admin/reset-database")
async def admin_reset_database(request: ResetRequest):
    """Reset all database data (ADMIN ONLY - requires password)"""
    try:
        # Verify admin password from environment
        ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "supply2024reset")
        
        if request.admin_password != ADMIN_PASSWORD:
            raise HTTPException(status_code=403, detail="Invalid admin password")
        
        # Get all logs to delete associated images
        try:
            logs = await prisma.logs.find_many()
        except Exception as db_error:
            print(f"Database query failed: {db_error}")
            raise HTTPException(status_code=500, detail="Database connection error")
        
        # Delete all image files from uploads folder
        deleted_files = 0
        for log in logs:
            if log.imagePath:
                filename = log.imagePath.split("/")[-1]
                file_path = os.path.join(UPLOAD_DIR, filename)
                if os.path.exists(file_path):
                    os.remove(file_path)
                    deleted_files += 1
        
        # Delete all logs from database
        try:
            deleted_logs = await prisma.logs.delete_many()
        except Exception as db_error:
            print(f"Database delete failed: {db_error}")
            raise HTTPException(status_code=500, detail="Database connection error")
        
        print(f"DATABASE RESET: {deleted_logs} logs deleted, {deleted_files} files removed")
        
        return {
            "status": "success",
            "message": "Database reset successfully",
            "deleted_logs": deleted_logs,
            "deleted_files": deleted_files
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Reset database error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to reset database: {str(e)}")

# ==============================
# API KEY MANAGEMENT (BYOK)
# ==============================

class ApiKeyRequest(BaseModel):
    apiKey: str
    provider: str = "openai"  # openai for chatbot
    isActive: bool = True  # Toggle state

@app.get("/api/user/apikey")
async def get_user_api_key(provider: str = "openai", authorization: str = Header(None)):
    """Get user's API key for specific provider (openai or ocrspace)"""
    try:
        email = get_user_email_from_token(authorization)
        
        api_key_record = await prisma.apikey.find_first(
            where={
                "userId": email,
                "provider": provider
            }
        )
        
        if not api_key_record:
            return {
                "hasApiKey": False,
                "provider": None,
                "createdAt": None,
                "updatedAt": None
            }
        
        # Return the full encrypted key. The frontend will handle masking for display.
        return {
            "hasApiKey": True,
            "apiKey": api_key_record.apiKey,
            "provider": api_key_record.provider,
            "isActive": api_key_record.isActive,
            "createdAt": api_key_record.createdAt.isoformat(),
            "updatedAt": api_key_record.updatedAt.isoformat()
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Get API key error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to get API key: {str(e)}")

@app.post("/api/user/apikey")
async def save_user_api_key(request: ApiKeyRequest, authorization: str = Header(None)):
    """Save or update user's API key"""
    try:
        email = get_user_email_from_token(authorization)
        
        # Check if API key already exists for this provider
        existing_key = await prisma.apikey.find_first(
            where={
                "userId": email,
                "provider": request.provider
            }
        )
        
        if existing_key:
            # Update existing key
            updated_key = await prisma.apikey.update(
                where={"id": existing_key.id},
                data={
                    "apiKey": request.apiKey,
                    "provider": request.provider,
                    "isActive": request.isActive
                }
            )
            return {
                "message": "API key updated successfully",
                "provider": updated_key.provider,
                "updatedAt": updated_key.updatedAt.isoformat()
            }
        else:
            # Create new key
            new_key = await prisma.apikey.create(
                data={
                    "userId": email,
                    "apiKey": request.apiKey,
                    "provider": request.provider,
                    "isActive": request.isActive
                }
            )
            return {
                "message": "API key saved successfully",
                "provider": new_key.provider,
                "createdAt": new_key.createdAt.isoformat()
            }
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Save API key error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to save API key: {str(e)}")

@app.delete("/api/user/apikey")
async def delete_user_api_key(provider: str = "openai", authorization: str = Header(None)):
    """Delete user's API key for specific provider"""
    try:
        email = get_user_email_from_token(authorization)
        
        # Check if API key exists
        existing_key = await prisma.apikey.find_first(
            where={
                "userId": email,
                "provider": provider
            }
        )
        
        if not existing_key:
            raise HTTPException(status_code=404, detail="API key not found")
        
        # Delete the key
        await prisma.apikey.delete(
            where={"id": existing_key.id}
        )
        
        return {
            "message": "API key deleted successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f" Delete API key error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to delete API key: {str(e)}")

@app.patch("/api/user/apikey/toggle")
async def toggle_api_key_status(isActive: bool, provider: str = "openai", authorization: str = Header(None)):
    """Toggle API key active status without modifying the key itself"""
    try:
        email = get_user_email_from_token(authorization)
        
        # Check if API key exists
        existing_key = await prisma.apikey.find_first(
            where={
                "userId": email,
                "provider": provider
            }
        )
        
        if not existing_key:
            raise HTTPException(status_code=404, detail="API key not found")
        
        # Update only the isActive field
        updated_key = await prisma.apikey.update(
            where={"id": existing_key.id},
            data={"isActive": isActive}
        )
        
        status_text = "diaktifkan" if isActive else "dinonaktifkan"
        return {
            "message": f"BYOK berhasil {status_text}",
            "isActive": updated_key.isActive
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Toggle API key error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to toggle API key: {str(e)}")

async def get_user_api_key_internal(email: str, provider: str = "openai") -> str | None:
    """Internal helper to get user's API key for specific provider - only if active"""
    try:
        api_key_record = await prisma.apikey.find_first(
            where={
                "userId": email,
                "provider": provider
            }
        )
        if api_key_record and api_key_record.isActive:
            return api_key_record.apiKey
        return None
    except Exception as e:
        print(f"Error getting user API key: {str(e)}")
        return None

async def get_user_ocr_api_key_internal(email: str) -> str | None:
    """Internal helper to get user's OCR API key - only if active"""
    return await get_user_api_key_internal(email, provider="ocrspace")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)


